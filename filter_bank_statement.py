import tkinter as tk
from tkinter import filedialog
import pandas as pd
import os
from openpyxl import load_workbook

def filter_and_copy():
    # Ask user to upload the input Excel file
    input_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])

    # Read the input Excel file
    df = pd.read_excel(input_file_path, header=None)  # Avoid using header to consider the 16th row as data
    print("Original DataFrame:")
    print(df)



    # Find the column index where "particulars" is located in the 16th row
    column_index = df.iloc[15].eq('PARTICULARS').idxmax()

    if column_index is not None:
        # Filter values in 'particulars' column
        edc_df = df[df.iloc[:, column_index].str.contains('EDC', case=False, na=False)]
        # Filter values in 'PARTICULARS' column containing the words 'salary' or 'ADC'
        salary_adc_df = df[df.iloc[:, column_index].str.contains('salary|ADC', case=False, na=False)]

        # Assign specific column names to the DataFrames
        edc_df.columns = ['']
        salary_adc_df.columns = ['']

        # Remove leading spaces and convert 'DR' and 'CR' columns to numeric format
        edc_df['Debit'] = pd.to_numeric(edc_df['Debit'].str.strip(), errors='coerce')
        edc_df['Credit'] = pd.to_numeric(edc_df['Credit'].str.strip(), errors='coerce')

        salary_adc_df['Debit'] = pd.to_numeric(salary_adc_df['Debit'].str.strip(), errors='coerce')
        salary_adc_df['Credit'] = pd.to_numeric(salary_adc_df['Credit'].str.strip(), errors='coerce')
        
        # filtered_df = df[df.iloc[:, column_index].str.contains('EDC', case=False, na=False)]
        # filtered_df.columns = ['S.NO', 'Transaction Date', 'CHQNO', 'PARTICULARS', 'Debit', 'Credit', 'Balance', 'SOL']
        # filtered_df['Debit'] = pd.to_numeric(filtered_df['Debit'].str.strip(), errors='coerce')
        # filtered_df['Credit'] = pd.to_numeric(filtered_df['Credit'].str.strip(), errors='coerce')
        


        print("Filtered DataFrame:")
        print(salary_adc_df)


        # Set the output file path (draft.xlsx on the desktop)
        # desktop_path = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')
        desktop_path = r""
        output_file_path = os.path.join(desktop_path, '')

        # Check if 'draft.xlsx' file already exists
        if os.path.isfile(output_file_path):
            # book = load_workbook(output_file_path)
            existing_data_dpay = pd.read_excel(output_file_path, sheet_name="DPAY",header=0)
            existing_data_Salary = pd.read_excel(output_file_path, sheet_name="Salary",header=0) 
            print("Data",existing_data_Salary.columns)
            # sheet_names = book.sheetnames
            # print("sheet names",sheet_names)
            # if 'DPAY' not in sheet_names:
            #     book.create_sheet('DPAY')
            # if 'Salary' not in sheet_names:
            #     book.create_sheet('Salary')
            # existing_edc_df = book['DPAY']
            # existing_salary_df = book['Salary']

            new_entries_edc = edc_df[~edc_df['Transaction Date'].isin(existing_data_dpay['Transaction Date'])]
            new_entries_salary = salary_adc_df[~salary_adc_df['Transaction Date'].isin(existing_data_Salary['Transaction Date'])]





            # Extract column names from the worksheet
            # dpay_columns = [cell.value for cell in existing_edc_df[1]]
            # salary_columns = [cell.value for cell in existing_salary_df[1]]

            # print("Columns in DPAY sheet:", dpay_columns)
            # print("Columns in Salary sheet:", salary_columns)


            # print("columns,",existing_edc_df)
            # print("type,",type(existing_edc_df))
        
            # Read existing data from 'draft.xlsx'
            # existing_df = pd.read_excel(output_file_path, header=0)
            # new_entries = filtered_df[~filtered_df['Transaction Date'].isin(existing_df['Transaction Date'])]
            # existing_edc_df = pd.read_excel(output_file_path, sheet_name='DPAY', header=0,engine='openpyxl')
            # existing_salary_df = pd.read_excel(output_file_path, sheet_name='Salary', header=0,engine='openpyxl')

             



            if not new_entries_edc.empty:

                # Append the filtered data to the existing data
                # filtered_df = filtered_df.drop(['S.No','CHQNO','SOL'],axis =1)
                columns_to_drop = ['']
                new_entries_edc = new_entries_edc.drop(columns=[col for col in columns_to_drop if col in new_entries_edc], errors='ignore')
                updated_df_edc = pd.concat([existing_data_dpay, new_entries_edc], ignore_index=True)
                # Write the updated data back to 'draft.xlsx'
                with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                    updated_df_edc.to_excel(writer,sheet_name='DPAY', index=False, header=True)
                # updated_df_edc.to_excel(existing_data_dpay, index=False, header=True)
                    result_label.config(text="Filter applied and data copied successfully to DPAY")
            else:
                result_label.config(text="No new entries found to append in DPAY.")

            if not new_entries_salary.empty:

                # Append the filtered data to the existing data
                # filtered_df = filtered_df.drop(['S.No','CHQNO','SOL'],axis =1)
                columns_to_drop = ['']
                new_entries_salary = new_entries_salary.drop(columns=[col for col in columns_to_drop if col in new_entries_salary], errors='ignore')
                updated_df_salary = pd.concat([existing_data_Salary, new_entries_salary], ignore_index=True)
                # Write the updated data back to 'draft.xlsx'
                # updated_df_salary.to_excel(existing_data_Salary, index=False, header=True)
                with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
                    updated_df_salary.to_excel(writer,sheet_name='Salary', index=False, header=True)
                    result_label.config(text="Filter applied and data copied successfully to Salary")
                
            else:
                result_label.config(text="No new entries found to append in Salary.")


        else:
            result_label.config(text="")

# Create the main window
root = tk.Tk()
root.title("Excel Filter and Copy")

# Create and pack UI elements
filter_button = tk.Button(root, text="Filter and Copy", command=filter_and_copy)
filter_button.pack(pady=20)

result_label = tk.Label(root, text="")
result_label.pack()

# Start the main event loop
root.mainloop()
